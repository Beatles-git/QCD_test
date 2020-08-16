from openpyxl import load_workbook
def duquceshiyongli(workbook,sheet):
    wb=load_workbook(workbook)
    sheet=wb[sheet]
    all_case=[]
    for i in range(sheet.max_row):
        case=[]
        for j in range(sheet.max_column):
            value=sheet.cell(row=i+1,column=j+1).value
            case.append(value)
        all_case.append(case)
    return all_case
if __name__ == '__main__':

    all_case=duquceshiyongli('test_case.xlsx','recharge')