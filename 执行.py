from QCD_test.读取表格数据 import duquceshiyongli
from QCD_test.请求 import qingqiu
from openpyxl import load_workbook
all_case=duquceshiyongli('test_case.xlsx','register')
global token
for i in range(1,len(all_case)):
    print(all_case[i])
    url=all_case[i][6]
    param = eval(all_case[i][7])
    header= eval(all_case[i][3])
    if 'register' in all_case[i][6]:
        response=qingqiu(url,param,header)
    elif 'login' in all_case[i][6]:
        response=qingqiu(url,param,header)
        token=response['data']['token_info']['token']
    else:
        header['Authorization']='Bearer '+token
        response = qingqiu(url, param, header)
    print(response)
    wb = load_workbook('test_case.xlsx')
    sheet=wb['register']
    sheet.cell(row=i+2,column=11).value=str(response)
    actual={'code':response['code'],'msg':response['msg']}
    expected=eval(all_case[i][8])
    if actual==expected:
        print('测试通过')
        sheet.cell(row=i+2,column=12).value='pass'
    else:
        print('测试不通过')
        sheet.cell(row=i+2,column=12).value='fail'

    wb.save('test_case.xlsx')

